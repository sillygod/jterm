"""SQL service for database connections and query execution.

This service provides database connectivity for SQLite and PostgreSQL,
query execution, schema introspection, and result export functionality.
"""

import aiosqlite
import asyncpg
import time
from pathlib import Path
from typing import Optional, List, Dict, Any, Union
import openpyxl
from openpyxl.styles import Font, PatternFill
import logging

from ..models.database import (
    DatabaseConnection,
    DatabaseType,
    TableSchema,
    ColumnSchema,
    QueryResult,
    QueryHistory
)

logger = logging.getLogger(__name__)


class SQLService:
    """Service for SQL database operations."""

    def __init__(self):
        """Initialize SQL service."""
        self._active_connections: Dict[str, Union[aiosqlite.Connection, asyncpg.Connection]] = {}

    async def connect_database(
        self,
        db_connection: DatabaseConnection
    ) -> DatabaseConnection:
        """Connect to a database and populate schema metadata.

        Args:
            db_connection: Database connection configuration

        Returns:
            Updated DatabaseConnection with connection status
        """
        try:
            if db_connection.db_type == DatabaseType.SQLITE:
                conn = await self._connect_sqlite(db_connection.connection_string)
            elif db_connection.db_type == DatabaseType.POSTGRESQL:
                conn = await self._connect_postgres(db_connection.connection_string)
            else:
                raise ValueError(f"Unsupported database type: {db_connection.db_type}")

            # Store connection
            conn_id = self._get_connection_id(db_connection)
            self._active_connections[conn_id] = conn

            # Introspect schema
            db_connection.tables = await self.introspect_schema(db_connection)
            db_connection.is_connected = True
            db_connection.error_message = None

            logger.info(f"Connected to {db_connection.db_type.value} database: {db_connection.display_name}")
            return db_connection

        except Exception as e:
            logger.error(f"Failed to connect to database: {str(e)}")
            db_connection.is_connected = False
            db_connection.error_message = str(e)
            return db_connection

    async def _connect_sqlite(self, connection_string: str) -> aiosqlite.Connection:
        """Connect to SQLite database.

        Args:
            connection_string: SQLite connection string (sqlite:///path/to/db)

        Returns:
            SQLite connection
        """
        # Extract path from connection string
        db_path = connection_string.replace("sqlite:///", "")

        # Check if file exists
        if not Path(db_path).exists():
            raise FileNotFoundError(f"SQLite database not found: {db_path}")

        conn = await aiosqlite.connect(db_path)
        # Enable foreign keys
        await conn.execute("PRAGMA foreign_keys = ON")
        return conn

    async def _connect_postgres(self, connection_string: str) -> asyncpg.Connection:
        """Connect to PostgreSQL database.

        Args:
            connection_string: PostgreSQL DSN (postgresql://user:pass@host:port/db)

        Returns:
            PostgreSQL connection
        """
        conn = await asyncpg.connect(connection_string)
        return conn

    async def execute_query(
        self,
        db_connection: DatabaseConnection,
        query: str,
        offset: int = 0,
        limit: int = 1000
    ) -> QueryResult:
        """Execute SQL query and return results.

        Args:
            db_connection: Database connection
            query: SQL query to execute
            offset: Row offset for pagination
            limit: Maximum rows to return (1-10000)

        Returns:
            Query execution result
        """
        # Validate limit
        if limit < 1 or limit > 10000:
            raise ValueError("Limit must be between 1 and 10000")

        conn_id = self._get_connection_id(db_connection)
        conn = self._active_connections.get(conn_id)

        if not conn:
            raise RuntimeError("Database not connected. Call connect_database() first.")

        # Add LIMIT and OFFSET to query if not present
        query_with_pagination = self._add_pagination(query, offset, limit)

        start_time = time.perf_counter()

        try:
            if db_connection.db_type == DatabaseType.SQLITE:
                result = await self._execute_sqlite_query(conn, query_with_pagination, offset, limit)
            elif db_connection.db_type == DatabaseType.POSTGRESQL:
                result = await self._execute_postgres_query(conn, query_with_pagination, offset, limit)
            else:
                raise ValueError(f"Unsupported database type: {db_connection.db_type}")

            execution_time_ms = (time.perf_counter() - start_time) * 1000
            result.execution_time_ms = execution_time_ms
            result.query = query

            logger.info(f"Query executed in {execution_time_ms:.2f}ms, returned {result.row_count} rows")
            return result

        except Exception as e:
            execution_time_ms = (time.perf_counter() - start_time) * 1000
            logger.error(f"Query failed after {execution_time_ms:.2f}ms: {str(e)}")
            raise

    async def _execute_sqlite_query(
        self,
        conn: aiosqlite.Connection,
        query: str,
        offset: int,
        limit: int
    ) -> QueryResult:
        """Execute query on SQLite database."""
        async with conn.execute(query) as cursor:
            columns = [desc[0] for desc in cursor.description] if cursor.description else []
            rows = await cursor.fetchall()

            # Check if there are more rows
            has_more = len(rows) == limit

            return QueryResult(
                columns=columns,
                rows=[list(row) for row in rows],
                row_count=len(rows),
                execution_time_ms=0,  # Will be set by caller
                offset=offset,
                limit=limit,
                has_more=has_more
            )

    async def _execute_postgres_query(
        self,
        conn: asyncpg.Connection,
        query: str,
        offset: int,
        limit: int
    ) -> QueryResult:
        """Execute query on PostgreSQL database."""
        rows = await conn.fetch(query)

        if rows:
            columns = list(rows[0].keys())
            row_data = [[row[col] for col in columns] for row in rows]
        else:
            columns = []
            row_data = []

        # Check if there are more rows
        has_more = len(rows) == limit

        return QueryResult(
            columns=columns,
            rows=row_data,
            row_count=len(rows),
            execution_time_ms=0,  # Will be set by caller
            offset=offset,
            limit=limit,
            has_more=has_more
        )

    async def introspect_schema(
        self,
        db_connection: DatabaseConnection
    ) -> List[TableSchema]:
        """Introspect database schema to get table and column information.

        Args:
            db_connection: Database connection

        Returns:
            List of table schemas
        """
        conn_id = self._get_connection_id(db_connection)
        conn = self._active_connections.get(conn_id)

        if not conn:
            raise RuntimeError("Database not connected. Call connect_database() first.")

        if db_connection.db_type == DatabaseType.SQLITE:
            return await self._introspect_sqlite_schema(conn)
        elif db_connection.db_type == DatabaseType.POSTGRESQL:
            return await self._introspect_postgres_schema(conn)
        else:
            raise ValueError(f"Unsupported database type: {db_connection.db_type}")

    async def _introspect_sqlite_schema(
        self,
        conn: aiosqlite.Connection
    ) -> List[TableSchema]:
        """Introspect SQLite database schema."""
        tables = []

        # Get all table names
        async with conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ) as cursor:
            table_names = [row[0] for row in await cursor.fetchall()]

        # For each table, get column information
        for table_name in table_names:
            columns = []

            # Get column info using PRAGMA
            async with conn.execute(f"PRAGMA table_info({table_name})") as cursor:
                for row in await cursor.fetchall():
                    column = ColumnSchema(
                        name=row[1],
                        data_type=row[2],
                        nullable=not bool(row[3]),
                        primary_key=bool(row[5]),
                        default_value=row[4]
                    )
                    columns.append(column)

            # Get row count
            async with conn.execute(f"SELECT COUNT(*) FROM {table_name}") as cursor:
                row_count = (await cursor.fetchone())[0]

            # Get indexes
            async with conn.execute(f"PRAGMA index_list({table_name})") as cursor:
                indexes = [row[1] for row in await cursor.fetchall()]

            table = TableSchema(
                name=table_name,
                columns=columns,
                row_count=row_count,
                indexes=indexes
            )
            tables.append(table)

        return tables

    async def _introspect_postgres_schema(
        self,
        conn: asyncpg.Connection
    ) -> List[TableSchema]:
        """Introspect PostgreSQL database schema."""
        tables = []

        # Get all table names from public schema
        table_rows = await conn.fetch("""
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema = 'public'
            ORDER BY table_name
        """)

        for table_row in table_rows:
            table_name = table_row['table_name']
            columns = []

            # Get column information
            column_rows = await conn.fetch("""
                SELECT column_name, data_type, is_nullable, column_default
                FROM information_schema.columns
                WHERE table_schema = 'public' AND table_name = $1
                ORDER BY ordinal_position
            """, table_name)

            # Get primary keys
            pk_rows = await conn.fetch("""
                SELECT a.attname
                FROM pg_index i
                JOIN pg_attribute a ON a.attrelid = i.indrelid AND a.attnum = ANY(i.indkey)
                WHERE i.indrelid = $1::regclass AND i.indisprimary
            """, table_name)

            primary_keys = {row['attname'] for row in pk_rows}

            for col_row in column_rows:
                column = ColumnSchema(
                    name=col_row['column_name'],
                    data_type=col_row['data_type'],
                    nullable=col_row['is_nullable'] == 'YES',
                    primary_key=col_row['column_name'] in primary_keys,
                    default_value=col_row['column_default']
                )
                columns.append(column)

            # Get row count (estimated for performance)
            count_result = await conn.fetchval(f"SELECT COUNT(*) FROM {table_name}")

            # Get indexes
            index_rows = await conn.fetch("""
                SELECT indexname
                FROM pg_indexes
                WHERE schemaname = 'public' AND tablename = $1
            """, table_name)

            indexes = [row['indexname'] for row in index_rows]

            table = TableSchema(
                name=table_name,
                columns=columns,
                row_count=count_result,
                indexes=indexes
            )
            tables.append(table)

        return tables

    async def export_results(
        self,
        result: QueryResult,
        format: str = "csv",
        file_path: Optional[str] = None
    ) -> str:
        """Export query results to file.

        Args:
            result: Query result to export
            format: Export format (csv, json, xlsx)
            file_path: Optional file path (if None, returns string)

        Returns:
            Exported data as string or file path
        """
        if format == "csv":
            data = result.to_csv()
        elif format == "json":
            data = result.to_json()
        elif format == "xlsx":
            data = await self._export_to_excel(result, file_path)
            return data  # Returns file path
        else:
            raise ValueError(f"Unsupported export format: {format}")

        if file_path:
            with open(file_path, 'w') as f:
                f.write(data)
            return file_path

        return data

    async def _export_to_excel(
        self,
        result: QueryResult,
        file_path: Optional[str]
    ) -> str:
        """Export query results to Excel file.

        Args:
            result: Query result to export
            file_path: File path for Excel file

        Returns:
            File path
        """
        if not file_path:
            file_path = f"query_result_{int(time.time())}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Query Results"

        # Write headers with styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, column_name in enumerate(result.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font

        # Write data rows
        for row_idx, row in enumerate(result.rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                # Convert to string for Excel compatibility
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)
        return file_path

    async def close_connection(self, db_connection: DatabaseConnection):
        """Close database connection.

        Args:
            db_connection: Database connection to close
        """
        conn_id = self._get_connection_id(db_connection)
        conn = self._active_connections.get(conn_id)

        if conn:
            if db_connection.db_type == DatabaseType.SQLITE:
                await conn.close()
            elif db_connection.db_type == DatabaseType.POSTGRESQL:
                await conn.close()

            del self._active_connections[conn_id]
            logger.info(f"Closed connection to {db_connection.display_name}")

    def _get_connection_id(self, db_connection: DatabaseConnection) -> str:
        """Generate unique connection ID."""
        return f"{db_connection.db_type.value}:{db_connection.connection_string}"

    def _add_pagination(self, query: str, offset: int, limit: int) -> str:
        """Add LIMIT and OFFSET to query if not already present.

        Args:
            query: Original SQL query
            offset: Row offset
            limit: Row limit

        Returns:
            Query with pagination
        """
        query_upper = query.upper().strip()

        # Check if LIMIT already exists
        if 'LIMIT' in query_upper:
            return query

        # Add LIMIT and OFFSET
        if offset > 0:
            return f"{query.rstrip(';')} LIMIT {limit} OFFSET {offset}"
        else:
            return f"{query.rstrip(';')} LIMIT {limit}"
